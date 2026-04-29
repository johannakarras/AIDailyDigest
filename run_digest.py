#!/usr/bin/env python3
"""AI Daily Digest — main orchestrator.

Usage:
    python run_digest.py           # normal run
    python run_digest.py --dry-run # fetch + filter but don't save or write HTML
"""

import csv
import json
import os
import re
import subprocess
import sys
import tempfile
from datetime import datetime, timezone, timedelta

import anthropic
from dotenv import load_dotenv

from pipeline.dedup import deduplicate_candidates, fetch_missing_abstracts, load_seen_ids
from pipeline.filter import filter_for_novelty
from pipeline.formatter import format_papers
from pipeline.rater import rate_papers
from sources.affiliations import fetch_affiliations, fetch_affiliations_from_pdf
from sources.arxiv_source import fetch_arxiv_papers
from web.generate import generate_html

DIGESTS_PATH = "data/digests.json"
HTML_PATH = "web/index.html"
MAX_WEEKLY = 5


def _get_week_key(d) -> str:
    """Return ISO week key string (e.g. '2026-W18') for a given date."""
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def _migrate_to_weekly(digests: dict) -> tuple[dict, bool]:
    """Convert any per-day entries (YYYY-MM-DD) to per-week entries (YYYY-Www).
    Returns (migrated_dict, changed)."""
    has_daily = any(re.match(r'^\d{4}-\d{2}-\d{2}$', k) for k in digests)
    if not has_daily:
        return digests, False

    new_digests = {}
    for key, papers in digests.items():
        if re.match(r'^\d{4}-\d{2}-\d{2}$', key):
            d = datetime.strptime(key, "%Y-%m-%d").date()
            week_key = _get_week_key(d)
            if week_key not in new_digests:
                new_digests[week_key] = []
            new_digests[week_key].extend(papers)
        else:
            new_digests[key] = papers

    # De-duplicate and keep top MAX_WEEKLY per week
    for k in new_digests:
        papers = new_digests[k]
        seen = {}
        for p in papers:
            if p.get("id") and p["id"] not in seen:
                seen[p["id"]] = p
        papers = list(seen.values())
        if len(papers) > MAX_WEEKLY:
            papers.sort(
                key=lambda p: (p.get("stars", 0), p.get("rating_total", 0.0)),
                reverse=True,
            )
            papers = papers[:MAX_WEEKLY]
        new_digests[k] = papers

    return new_digests, True


def main(dry_run: bool = False) -> None:
    load_dotenv()

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set. Add it to .env or your environment.")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    today = datetime.now(timezone.utc).date()
    week_key = _get_week_key(today)
    print(f"[main] Running digest for {today} (week {week_key}){' (dry run)' if dry_run else ''}")

    digests = _load_digests(DIGESTS_PATH)
    digests, migrated = _migrate_to_weekly(digests)
    if migrated:
        print("[main] Migrated daily entries to weekly format")

    week_papers = digests.get(week_key, [])
    seen_ids = load_seen_ids(digests)
    print(f"[main] Week {week_key}: {len(week_papers)} paper(s) so far | {len(seen_ids)} total seen")

    print("[main] Fetching papers from sources...")
    candidates = fetch_arxiv_papers(days_back=7, max_results=100)
    print(f"[sources] arxiv={len(candidates)}")
    candidates = deduplicate_candidates(candidates, seen_ids)
    print(f"[dedup] {len(candidates)} unique new candidates")

    candidates = fetch_missing_abstracts(candidates)
    candidates = [p for p in candidates if p.get("abstract")]
    print(f"[dedup] {len(candidates)} candidates with abstracts")

    cutoff = (today - timedelta(days=14)).isoformat()
    before = len(candidates)
    candidates = [p for p in candidates if p.get("submitted_date", "9999") >= cutoff]
    if len(candidates) < before:
        print(f"[dedup] Dropped {before - len(candidates)} paper(s) older than 14 days (cutoff {cutoff})")

    if not candidates:
        print("[main] No new candidates. Nothing to do.")
        if migrated and not dry_run:
            _save_digests(digests, DIGESTS_PATH)
            generate_html(digests, output_path=HTML_PATH)
        return

    print("[main] Fetching author affiliations...")
    aff_map = fetch_affiliations([p["id"] for p in candidates])
    for p in candidates:
        p["affiliations"] = aff_map.get(p["id"], [])

    print("[main] Filtering for novelty...")
    passing, rejected = filter_for_novelty(candidates, client)
    print(f"[filter] {len(passing)}/{len(candidates)} passed novelty filter")

    if rejected and not dry_run:
        _save_rejected_csv(rejected, today.isoformat())
        print(f"[main] Saved {len(rejected)} rejected papers to rejected/{today.isoformat()}.csv")

    if not passing:
        print("[main] No novel papers found this run.")
        if migrated and not dry_run:
            _save_digests(digests, DIGESTS_PATH)
            generate_html(digests, output_path=HTML_PATH)
        return

    # Rate ALL passing papers cheaply (haiku) — used to rank against existing week papers
    print(f"[main] Rating {len(passing)} passing paper(s)...")
    rate_papers(passing, client)

    # Merge new passing papers with existing week papers; keep top MAX_WEEKLY by (stars, total)
    existing_ids = {p["id"] for p in week_papers}
    new_passing = [p for p in passing if p["id"] not in existing_ids]
    combined = week_papers + new_passing
    combined.sort(
        key=lambda p: (p.get("stars", 0), p.get("rating_total", 0.0)),
        reverse=True,
    )
    top_n = combined[:MAX_WEEKLY]

    # Log any swaps
    top_ids = {p["id"] for p in top_n}
    dropped = [p for p in week_papers if p["id"] not in top_ids]
    added = [p for p in top_n if p["id"] not in existing_ids]
    if dropped:
        print(f"[main] Dropped {len(dropped)} paper(s) from week: {[p['title'][:50] for p in dropped]}")
    if added:
        print(f"[main] Added {len(added)} new paper(s) to week: {[p['title'][:50] for p in added]}")
    if not dropped and not added:
        print("[main] No changes to this week's digest.")
        if migrated and not dry_run:
            _save_digests(digests, DIGESTS_PATH)
            generate_html(digests, output_path=HTML_PATH)
        return

    # Format only new papers that made the cut (existing week papers already have description etc.)
    need_format = [p for p in top_n if not p.get("description")]
    if need_format:
        print(f"[main] Fetching PDF affiliations for {len(need_format)} new featured paper(s)...")
        for p in need_format:
            if not p.get("affiliations"):
                affs = fetch_affiliations_from_pdf(p["id"], client)
                if affs:
                    p["affiliations"] = affs
                    print(f"[affiliations] PDF: {p['id']} → {affs}")

        print(f"[main] Formatting {len(need_format)} new paper(s)...")
        format_papers(need_format, client, max_papers=len(need_format))
        # Remove any that failed formatting
        top_n = [p for p in top_n if p.get("description")]

    print(f"\n--- Week {week_key} ({len(top_n)} papers) ---")
    for p in sorted(top_n, key=lambda p: p.get("stars", 0), reverse=True):
        print(f"  {'⭐' * p.get('stars', 0)} {p['title']}")
    print()

    if dry_run:
        print("[dry-run] Skipping save and HTML generation.")
        return

    digests[week_key] = top_n
    _save_digests(digests, DIGESTS_PATH)
    print(f"[main] Saved week {week_key} with {len(top_n)} papers to {DIGESTS_PATH}")

    generate_html(digests, output_path=HTML_PATH)
    print(f"[main] Regenerated {HTML_PATH}")
    print(f"[main] Done. Open {HTML_PATH} in a browser.")

    _git_commit_and_push(week_key, len(top_n))


def _load_digests(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    with open(path) as f:
        return json.load(f)


def _save_rejected_csv(rejected: list[dict], date: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs("rejected", exist_ok=True)
    rows = []
    for paper in rejected:
        authors = paper.get("authors") or []
        affiliations = paper.get("affiliations") or []
        rows.append({
            "title": paper.get("title", ""),
            "authors": "; ".join(authors) if authors else "",
            "affiliations": "; ".join(affiliations) if affiliations else "",
            "rejection_reason": paper.get("filter_reason", ""),
            "url": paper.get("url", ""),
        })

    csv_path = f"rejected/{date}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["title", "authors", "affiliations", "rejection_reason", "url"])
        writer.writeheader()
        writer.writerows(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date

    headers = ["Title", "Authors", "Affiliations", "Rejection Reason", "URL"]
    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["title"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=2, value=row["authors"])
        ws.cell(row=row_idx, column=3, value=row["affiliations"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4, value=row["rejection_reason"]).alignment = Alignment(wrap_text=True)
        url_cell = ws.cell(row=row_idx, column=5, value=row["url"])
        if row["url"]:
            url_cell.hyperlink = row["url"]
            url_cell.font = Font(color="1155CC", underline="single")

    col_widths = [60, 35, 40, 55, 45]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    xlsx_path = f"rejected/{date}.xlsx"
    wb.save(xlsx_path)


def _git_commit_and_push(week_key: str, num_papers: int) -> None:
    try:
        subprocess.run(["git", "add", DIGESTS_PATH, HTML_PATH], check=True)
        msg = f"digest: {week_key} ({num_papers} paper{'s' if num_papers != 1 else ''})"
        result = subprocess.run(["git", "diff", "--cached", "--quiet"])
        if result.returncode == 0:
            print("[git] Nothing new to commit.")
            return
        subprocess.run(["git", "commit", "-m", msg], check=True)
        subprocess.run(["git", "push", "--set-upstream", "origin", "main"], check=True)
        print(f"[git] Pushed: {msg}")
    except subprocess.CalledProcessError as e:
        print(f"[git] Push failed: {e}")


def _save_digests(digests: dict, path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".json.tmp")
    try:
        with os.fdopen(fd, "w") as f:
            json.dump(digests, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    except Exception:
        os.unlink(tmp_path)
        raise


if __name__ == "__main__":
    dry_run = "--dry-run" in sys.argv
    main(dry_run=dry_run)
